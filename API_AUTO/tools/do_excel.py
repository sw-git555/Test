from openpyxl import load_workbook
from tools.project_path import test_case_path, base_path, case_config_path
from tools.read_config import ReadConfig
from tools.get_global import GetAdmin, GetTeacher, GetStudent
from tools.get_date import str_time_mdhms
from tools.handle_sqlite import db
from tools.do_regx import DoRegx


class DoExcel:

    def get_data(self, file_name=None):
        global new_data
        excel_row = eval(ReadConfig().get_config(case_config_path, 'Excel_row', 'excel_row'))
        if file_name is None:
            wb = load_workbook(test_case_path)
        else:
            wb = load_workbook(base_path + 'test_data' + file_name)
        mode = eval(ReadConfig.get_config(case_config_path, 'Mode', 'mode'))
        test_data = []
        for key in mode:
            # print(key)
            sheet = wb[key]
            if mode[key] == 'all':
                list_range = range(1, sheet.max_row)
            else:
                list_range = mode[key]
            for i in list_range:
                i = i + 1
                new_data = sheet.cell(i, excel_row['data']).value
                if sheet.cell(i, excel_row['data']).value.find('${admin}') != -1:
                    new_admin = getattr(GetAdmin, 'admin') + str_time_mdhms
                    setattr(GetAdmin, 'admin', new_admin)
                    new_data = DoRegx().do_regx(sheet.cell(i, excel_row['data']).value, GetAdmin)
                if sheet.cell(i, excel_row['data']).value.find('${student}') != -1:
                    new_student = getattr(GetStudent, 'student') + str_time_mdhms
                    setattr(GetStudent, 'student', new_student)
                    new_data = DoRegx().do_regx(sheet.cell(i, excel_row['data']).value, GetStudent)
                if sheet.cell(i, excel_row['data']).value.find('${teacher}') != -1:
                    new_student = getattr(GetTeacher, 'teacher') + str_time_mdhms
                    setattr(GetTeacher, 'teacher', new_student)
                    new_data = DoRegx().do_regx(sheet.cell(i, excel_row['data']).value, GetTeacher)

                row_data = {'caseid': sheet.cell(i, excel_row['caseid']).value,
                            'url': sheet.cell(i, excel_row['url']).value,
                            'data': eval(new_data),
                            'title': sheet.cell(i, excel_row['title']).value,
                            'method': sheet.cell(i, excel_row['method']).value,
                            'data_type': sheet.cell(i, excel_row['data_type']).value,
                            'expect': sheet.cell(i, excel_row['expect']).value,
                            'test_result': sheet.cell(i, excel_row['test_result']).value,
                            'test_sql': sheet.cell(i, excel_row['test_sql']).value,
                            'sheet_name': key
                            }
                test_data.append(row_data)
        return test_data

    def write_back(self, i, j, result_data, file_name=None, sheet_name=None):
        if file_name is None:
            wb = load_workbook(test_case_path)
        else:
            wb = load_workbook(base_path + 'test_data' + file_name)
        if sheet_name is None:
            sheet = wb['login']
        else:
            sheet = wb[sheet_name]
        sheet.cell(i, j).value = result_data
        if file_name is None:
            wb.save(test_case_path)
        else:
            wb.save(base_path + 'test_data' + file_name)

    def write_all_back(self, result_data, test_result_data, file_name=None):
        if file_name is None:
            wb = load_workbook(test_case_path)
        else:
            wb = load_workbook(base_path + 'test_data' + file_name)
        excel_row = eval(ReadConfig().get_config(case_config_path, 'Excel_row', 'excel_row'))
        mode = eval(ReadConfig.get_config(case_config_path, 'Mode', 'mode'))
        for key in mode:
            sheet = wb[key]
            if mode[key] == 'all':
                list_range = range(1, sheet.max_row)
            else:
                list_range = mode[key]
            k = 0
            for i in list_range:
                i = i + 1
                sheet.cell(i, excel_row['result']).value = result_data[k]
                sheet.cell(i, excel_row['test_result']).value = test_result_data[k]
                k = k + 1
        if file_name is None:
            wb.save(test_case_path)
        else:
            wb.save(base_path + 'test_data' + file_name)


if __name__ == '__main__':
    a = DoExcel().get_data()
    print(type(a))
    print(a)
    # data = [1, 2]
    # DoExcel().write_all_back(result_data=data)
    # DoExcel().write_back(2, 'a')
